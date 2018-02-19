import csv
import openpyxl
from openpyxl import load_workbook

wb = openpyxl.Workbook()
ws = wb.active

f = open('myexcel.csv')
reader = csv.reader(f, delimiter=',')
for row in reader:
    ws.append(row)
f.close()

wb.save('myexcel.xlsx')
wb2 = load_workbook('myexcel.xlsx')
ws = wb2.active
print(ws['A1'].value)
for row in ws.iter_rows(min_row=1,min_col=11, max_col=11, max_row=40):
    for cell in row:
        print(cell.value)


