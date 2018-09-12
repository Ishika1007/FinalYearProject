from html.parser import HTMLParser
from bs4 import BeautifulSoup
from time import sleep
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import csv
import xml.etree.ElementTree as ET
import re
import networkx as nx
import matplotlib.pyplot as plt
from collections import namedtuple

import csv
import openpyxl
from openpyxl import load_workbook
k = 1
fields = ['requestMethod','url','host','referer','origin','csrfToken','form','sensitiveParameters','statusCode','contentType',
         'outgoingUrl','server','x_frame_options','x_xss','content_length','comments']

wbnew = openpyxl.Workbook()
wbnew.save('new_excel.xlsx')
ws2 = wbnew.active
# for loop here..............
#url = input("Enter a website to extract the URL's from: ")
#x = True
def getParameters(hostname,urlname):
    host = hostname
    host = check_suffix(host)
    URL = urlname
    URL = check_suffix(URL)
    #requestMethod,url,host,referer,origin,csrfToken,form,sensitiveParameters,statusCode,contentType,
     #    server,x_frame_options,x_xss,content_length,comments
    outgoingUrl = []
    r = ''
    Request = namedtuple('Request', fields)
    Main = []
    req1 = requests.get(URL, allow_redirects = False)
    req = requests.Request('GET',URL, headers={'host':host})
    s = req.prepare()
    sessions = requests.Session()
    retry = Retry(connect=3, backoff_factor=0.5)
    adapter = HTTPAdapter(max_retries=retry)


    try:
        r = sessions.get(URL)
        url = s.path_url #url
        data = r.text
        requestMethod = r.request.method #request method
        statusCode = r.status_code
        '''server = r.headers['Server'] '''
        #print(server)
        soup = BeautifulSoup(data,'html.parser')
        form = False      # form tag
        max = ws2.max_row+1
        #print(r.history)
        #print("HREFs:")
        count=0
        val1 = URL[(len(host)-1):]
        for link in soup.find_all('a'):
            if link.get('href'):
                #print(link.get('href'))
                outgoingUrl.append(link.get('href'))
                a = link.get('href')
                a = check_path(a)
                if check_file_type(a):
                    ws2.cell(row=max, column=1).value = val1
                    ws2.cell(row=max, column=2).value = a
                    wbnew.save('new_excel.xlsx')
                    max+=1
        #print("SCRIPT SOURCE:")
        for link in soup.find_all('script'):
            if link.get('src'):
                #print(link.get('src'))
                outgoingUrl.append(link.get('src'))
                a = link.get('src')
                a = check_path(a)
                if check_file_type(a):
                    ws2.cell(row=max, column=1).value = val1
                    ws2.cell(row=max, column=2).value = a
                    wbnew.save('new_excel.xlsx')
                    max+= 1
        #print("IMAGE SOURCE:")
        for link in soup.find_all('img'):
            if link.get('src'):
                #print(link.get('src'))
                outgoingUrl.append(link.get('src'))
                a = link.get('src')
                a = check_path(a)
                if check_file_type(a):
                    ws2.cell(row=max, column=1).value = val1
                    ws2.cell(row=max, column=2).value = a
                    wbnew.save('new_excel.xlsx')
                    max+=1
        #print("FORM ELEEMENT:")
        for link in soup.find_all('form'):
            if link.get('action'):
                #print(link.get('action'))
                form = True
                outgoingUrl.append(link.get('action'))
                a = link.get('action')
                a = check_path(a)
                if check_file_type(a):
                    ws2.cell(row=max, column=1).value = val1
                    ws2.cell(row=max, column=2).value = a
                    wbnew.save('new_excel.xlsx')
                    count+=1
                    max+= 1
        for link in soup.find_all('meta'):
            if link.get('URL'):
                #print(link.get('URL'))
                outgoingUrl.append(link.get('URL'))
                a = link.get('URL')
                a = check_path(a)
                if check_file_type(a):
                    ws2.cell(row=max, column=1).value = val1
                    ws2.cell(row=max, column=2).value = a
                    wbnew.save('new_excel.xlsx')
                    max+= 1
         # LINK HREF LEFT..............



    except requests.exceptions.ConnectionError as exc:
        print(exc)
        sleep(5)

def check_file_type(a):
    print(type(a))
    if ".css" in a or ".png" in a or ".ico" in a or ".woff" in a or ".woff2" in a or ".gif" in a or ".txt" in a:
        return False
    else:
        return True


def check_path(a):
    if a[0]=='/':
        return a
    else:
        return "/"+a

def check_suffix(a):
    if a[-1]=='/':
        return a
    else:
        return a+"/"
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    f = open('myexcel.csv')
    reader = csv.reader(f)
    for row in reader:
        ws.append(row)
    f.close()

    wb.save('myexcel.xlsx')
    wb2 = load_workbook('myexcel.xlsx')
    ws = wb2.active
#    print(ws['A1'].value)
    max = ws.max_row
    print(max)
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=1, max_row=max):
        for cell in row:
            getParameters(ws['A1'].value, cell.value)


if __name__ == "__main__":
    main()