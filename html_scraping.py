from html.parser import HTMLParser
from bs4 import BeautifulSoup
from time import sleep
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import csv
import xml.etree.ElementTree as ET
import re
import networkx as nx
import matplotlib.pyplot as plt
from collections import namedtuple

import csv
import openpyxl
from openpyxl import load_workbook
k = 1
fields = ['requestMethod','url','host','referer','origin','csrfToken','form','sensitiveParameters','statusCode','contentType',
         'outgoingUrl','server','x_frame_options','x_xss','content_length','comments']

wbnew = openpyxl.Workbook()
wbnew.save('new_excel.xlsx')
ws2 = wbnew.active
# for loop here..............
#url = input("Enter a website to extract the URL's from: ")
#x = True
def getParameters(hostname,urlname):
    host = hostname
    URL = urlname
    #requestMethod,url,host,referer,origin,csrfToken,form,sensitiveParameters,statusCode,contentType,
     #    server,x_frame_options,x_xss,content_length,comments
    outgoingUrl = []
    r = ''
    Request = namedtuple('Request', fields)
    Main = []
    req1 = requests.get(URL, allow_redirects = False)
    req = requests.Request('GET',URL, headers={'host':host})
    s = req.prepare()
    sessions = requests.Session()
    retry = Retry(connect=3, backoff_factor=0.5)
    adapter = HTTPAdapter(max_retries=retry)


    try:
        r = sessions.get(URL)
        url = s.path_url #url
        data = r.text
        requestMethod = r.request.method #request method
        host = s.headers['host'] #host name
        statusCode = r.status_code
        server = r.headers['Server']
        #print(server)
        soup = BeautifulSoup(data)
        form = False      # form tag
        max = ws2.max_row+1
        print("msx")
        print(max)
        #print(r.history)
        #print("HREFs:")
        count=0
        for link in soup.find_all('a'):
            if link.get('href'):
                #print(link.get('href'))
                outgoingUrl.append(link.get('href'))
                ws2.cell(row=max, column=1).value = urlname[len(hostname):]
                ws2.cell(row=max, column=2).value = link.get('href')
                wbnew.save('new_excel.xlsx')
                max+=1
        #print("SCRIPT SOURCE:")
        for link in soup.find_all('script'):
            if link.get('src'):
                #print(link.get('src'))
                outgoingUrl.append(link.get('src'))
                ws2.cell(row=max, column=1).value = urlname[len(hostname):]
                ws2.cell(row=max, column=2).value = link.get('src')
                wbnew.save('new_excel.xlsx')
                max+= 1
        #print("IMAGE SOURCE:")
        for link in soup.find_all('img'):
            if link.get('src'):
                #print(link.get('src'))
                outgoingUrl.append(link.get('src'))
                ws2.cell(row=max, column=1).value = urlname[len(hostname):]
                ws2.cell(row=max, column=2).value = link.get('src')
                wbnew.save('new_excel.xlsx')
                max+=1
        #print("FORM ELEEMENT:")
        for link in soup.find_all('form'):
            if link.get('action'):
                #print(link.get('action'))
                form = True
                outgoingUrl.append(link.get('action'))
                ws2.cell(row=max, column=1).value = urlname[len(hostname):]
                ws2.cell(row=max, column=2).value = link.get('action')
                wbnew.save('new_excel.xlsx')
                count+=1
                max+= 1
        for link in soup.find_all('meta'):
            if link.get('URL'):
                #print(link.get('URL'))
                outgoingUrl.append(link.get('URL'))
                ws2.cell(row=max, column=1).value = urlname[len(hostname):]
                ws2.cell(row=max, column=2).value = link.get('URL')
                wbnew.save('new_excel.xlsx')
                max+= 1
         # LINK HREF LEFT..............
        print(URL)
        print(count)
        #print(hostname)
        print(len(outgoingUrl))


    except requests.exceptions.ConnectionError as exc:
        print(exc)
        sleep(5)

def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    f = open('myexcel.csv')
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)
    f.close()

    wb.save('myexcel.xlsx')
    wb2 = load_workbook('myexcel.xlsx')
    ws = wb2.active
#    print(ws['A1'].value)

    for row in ws.iter_rows(min_row=1, min_col=17, max_col=17, max_row=135):
        for cell in row:
            getParameters(ws['A2'].value, cell.value)


if __name__ == "__main__":
    main()