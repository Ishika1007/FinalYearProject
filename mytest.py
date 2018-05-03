
import csv
import xml.etree.ElementTree as ET
import re
import nltk
import networkx as nx
import scipy as sp
import matplotlib.pyplot as plt
from collections import namedtuple
import networkx as nx
from bs4 import BeautifulSoup
from itertools import groupby
from time import sleep
import numpy as np
import requests
from nltk.tokenize import word_tokenize
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import csv
from nltk.stem import PorterStemmer
import openpyxl
from openpyxl import load_workbook
wbnew = openpyxl.Workbook()
wbnew.save('Flipkart_new.xlsx')
ws2 = wbnew.active
wbnew2 = openpyxl.Workbook()
wbnew2.save('Flipkart_data.xlsx')
ws3 = wbnew2.active
third = {}
mydata,mydata2 = set(),set()
G=nx.DiGraph()
mycount = 0
myglobal = {}
ps = PorterStemmer()
form_dict = {}
mylocation = set()
myhost = {}
def parseXML(xmlfile):

    initialize_fields()
    # create element tree object
    tree = ET.parse(xmlfile)

    # get root element
    root = tree.getroot()

    # create empty list for news items
    data = []
    outgoingUrl = []

    max = ws2.max_row + 1
    port = "80"
    protocol = "http"
    for item in root.findall('./Request'):
        # empty news dictionary
        request = {}
        ownPath = ""
        URL,host, host1,contentType = "", "","",""
        statusCode=100
        count_session,count_pay=0,0
        check = True
        # iterate child elements of item
        for child in item:
            script=0
            location = None
            request[child.tag] = child.text
            if child.tag == "Port":
                port = child.text
            if child.tag == "Protocol":
                protocol = child.text
            if child.tag == "Hostname":
                host = child.text
                host = host.strip()
                host1 = protocol.strip()+"://"+host+":"+port.strip()
          #      print("host1"+host1)
            if child.tag == "Url":
                URL = child.text
                count_pay = 0
                count_session = 0
                x_powered_by = ""


               # print(URL)
                count_pay = string_text_processing(URL,'nlp')
                count_session = string_text_processing(URL,'session')
                server = ""#r.headers['Server']
                contentType = ""#r.headers['Content-Type']
                contentLength = ""#r.headers['content-length']
                URL = URL.strip()
                URL = URL[len(host1):]
           #     print("url"+URL)
                URL = check_both(URL)
                if not check_file_type(URL):
                    break
            if child.tag == "RequestData":
                if not child.text:
                    continue
                parameters,values = getParameters(child.text)
                count_pay += list_text_processing(parameters, 'nlp')
                count_session += list_text_processing(parameters, 'session')
                count_pay+=parameter_value_processing(values)
            if child.tag == "RequestHeader":
                referer = processRequestHeaderReferer(child.text,host)
                referer = check_parameters(referer)
                referer = check_both(referer)
                requestMethod = getRequestMethod(child.text)
               # print(requestMethod)
            if child.tag == "ResponseHeader":
                child.text = child.text.strip()
                statusCode = child.text[9:12]
                if statusCode[0]=='3':
                    location = getLocation(child.text,host,host1,URL)
                  #  print(location)
                append_new(max, URL, None, requestMethod, host, statusCode, server, referer, 0, contentType,
                           contentLength,
                           x_powered_by, count_pay, count_session,location,protocol)

            #if child.tag ==
            if child.tag == "ResponseData":
                if not child.text:
                    continue
                #count_pay = string_text_processing(child.text,'nlp')
               # count_session = sessionprocessing(child.text)
                soup = BeautifulSoup(child.text, 'html.parser')
                for link in soup.find_all('a'):
                    if link.get('href'):
                        a=link.get('href')
                        if not check_file_type(a):
                            continue
                        a = re.sub(r'\\', '', a)
                        a = re.sub(r'\"', '', a)
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,0,contentType,contentLength,x_powered_by,count_pay,count_session,location,protocol)
                        check = True
                        max += 1
                for link in soup.find_all('script'):
                    script = 1
                    if link.get('src'):
                        a=link.get('src')
                        if not check_file_type(a):
                            continue
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,0,contentType,contentLength,x_powered_by,count_pay,count_session,location,protocol)
                        check = True
                        max += 1
                for link in soup.find_all('form'):
                    if link.get('action'):
                        a=link.get('action')
                        if not check_file_type(a):
                            continue
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,1,contentType,contentLength,x_powered_by,count_pay,count_session,location,protocol)
                        check = True
                        max += 1
                for link in soup.find_all('meta'):
                    if link.get('URL'):
                        a=link.get('URL')
                        if not check_file_type(a):
                            continue
                        append_new(max, URL, a, requestMethod,host,statusCode,server, referer,0,contentType,contentLength,x_powered_by,count_pay,count_session,location,protocol)
                        check = True
                        max += 1
                if not check:
                    append_new(max, URL, None, requestMethod, host, statusCode, server, referer, 0, contentType, contentLength,
                    x_powered_by, count_pay, count_session,location,protocol)

            # special checking for namespace object content:media

        # append news dictionary to news items list

def list_text_processing(child,file):
 #   text_translated = re.sub(r'[^a-z]', ' ', child.lower())
 #   text_translated = word_tokenize(text_translated)
    fileObj  = open(file, "r")
    fileObj = list(fileObj)
    count = 0

    for i, j in enumerate(fileObj):
        fileObj[i] = j[:-1]

    for i,j in enumerate(child):
        child[i] = j.lower()
    for i in fileObj:
        for j in child:
            if i in j:
                count += 1
                print("list" + i + " " + j)
    return count

def string_text_processing(child,file):
    text_translated = re.sub(r'[^a-z]', ' ', child.lower())
    text_translated = word_tokenize(text_translated)
    fileObj = open(file, "r")
    fileObj = list(fileObj)
    count = 0
    for i, j in enumerate(fileObj):
        fileObj[i] = j[:-1]
    for i in fileObj:
        for j in text_translated:
            if i in j:
                count += 1
                print("string"+i+" "+ j)
    return count

def getParameters(child):
    para_list = re.findall(r'\w+\=\w+',child)
   # print(para_list)
    if len(para_list)>0:
        list1 = []
        list2 = []
        for i in para_list:
            j = i.find("=")
            word1 = i[:j]
            text = re.sub(r'[^a-z]', ' ', word1.lower())
            text = word_tokenize(text)
            for k in text:
                list1.append(k)
            word1 = i[j+1:]
            text = re.sub(r'[^a-z]', '', word1.lower())
            text = word_tokenize(text)
            for k in text:
                if k=="":
                    continue
                list2.append(k)
        return list1,list2
    else:
        return "",""

def parameter_value_processing(mylist):
    count=0
    for j in mylist:
        if passport(j) or adhar_card(j) or credit_card(j) or cheque_num(j) or social_security_num(j):
            count+=1
    return count
def getRequestMethod(child):
    method = re.findall(r'[^/]*/',child)
    method = method[0]
    method = method[:-2]
    return method.strip()
def processRequestHeaderReferer(child,hostname):
    referer = re.findall(r'Referer: http://', child)
    referer2 = re.findall(r'Referer: https://', child)
    if len(referer)==0 and len(referer2)==0:
        return ""
    elif len(referer2)==0:
        referer = referer[0] + hostname
        refererString = (re.findall(r'Referer: http://.*', child))
        if len(refererString) == 0:
            return ""
        refererString = refererString[0]
        return check_both(refererString[len(referer):])
    else:
        referer = referer2[0] + hostname
        refererString = (re.findall(r'Referer: https://.*', child))
        if len(refererString) == 0:
            return ""
        refererString = refererString[0]
        return check_both(refererString[len(referer):])


def getLocation(child,host,host1,URL):
    location = re.findall(r'Location:.*',child)
    if len(location)==0:
        return None
    location = location[0]
    loc = "Location: "
    mylocation = location[len(loc):]
    s1 = "http://"+host
    s2 = "https://"+host
    if s1 in location:
        ans = check_parameters(location[(len(loc)+len(s1)):])
        ans = check_both(ans)
    elif s2 in location:
        ans = check_parameters(location[(len(loc) + len(s2)):])
        ans = check_both(ans)
    else:
        ans = check_parameters(mylocation.strip())
        ans = check_both(ans)
    return ans
def initialize_fields():

    ws2.cell(row=1, column=1).value = "Source URL"
    ws2.cell(row=1, column=2).value = "Outgoing URL"
    ws3.cell(row=1, column=1).value = "Page:"
    ws3.cell(row=1, column=2).value = "Outgoing:"
    ws3.cell(row=1, column=3).value = "Incoming:"
    ws3.cell(row=1, column=4).value = "Host name:"
    ws3.cell(row=1, column=5).value = "Indegree Centrality"
    ws3.cell(row=1, column=6).value = "Outdegree Centrality"
    ws3.cell(row=1, column=7).value = "Closeness Centrality"
    ws3.cell(row=1, column=8).value = "Betweeness Centrality"
    ws3.cell(row=1, column=9).value = "EigenVector Centrality"
    ws3.cell(row=1, column=10).value = "PageRank Score"
    ws3.cell(row=1, column=11).value = "Payment words:"
    ws3.cell(row=1, column=12).value = "Session words:"
    ws3.cell(row=1, column=13).value = "Number of form tags:"
    ws3.cell(row=1, column=14).value = "Method:"
    ws3.cell(row=1, column=15).value = "Have Third-party connection"

def append_new(max,URL,a,requestMethod,Host,statusCode,server, referer,form,ct,cl,xpb,cnt1,cnt2,location,protocol):
    URL = check_parameters(URL)
    protocol = protocol.strip()
    myhost[URL]=Host
    if check_third_party(a,Host):
        third[URL] = 1
        return
    if URL not in third:
        third[URL] = 0
    if a:
        a = check_parameters(a)
        if Host in a:
            print(Host + " "+a+" "+protocol)
            if protocol in a:
                x = protocol+"://"+Host
                print(x)
                a = check_both(a[len(x):])
            else:
                a = check_both(a[len(Host):])
            print(a)
        a = check_both(check_parameters(a))
        G.add_node(a)
        G.add_node(URL)
        G.add_node(referer)
        G.add_edge(referer,URL)
        G.add_edge(URL, a)
    else:
        G.add_node(URL)
        G.add_node(referer)
        G.add_edge(referer, URL)
    if location:
        mylocation.add((URL,location))

   #     G.add_node(location)
    #    G.add_edge(URL,location)
    mydata2.add((URL,a))
    mydata2.add((referer,URL))
    mydata.add((URL,a,requestMethod,Host,statusCode,server, form,ct,cl,cnt1,cnt2))

def check_parameters(link):
    if "?" in link:
        myindex = link.find("?")
        return check_suffix(link[:myindex])
    else:
        return link

def check_file_type(a):
    if ".css" in a or ".png" in a or ".ico" in a or ".woff" in a or ".woff2" in a or ".gif" in a or ".txt" in a or ".jpg" in a:
        return False
    else:
        return True
def check_path(a):
    if not a:
        return "/"
    if a[0]=="/":
        return a
    else:
        return "/"+a
def check_suffix(a):
    if not a:
        return "/"
    if a[-1]=='/':
        return a
    else:
        return a+"/"
def check_both(a):
    a = check_path(a)
    a = check_suffix(a)
    return a
#credit-card regex
def count_consecutive(a):
    length = 0
    for _, g in groupby(a):
        length = max(length,len(list(g)))
    return length

def credit_card(a):
    pattern = re.compile(r'(?:\d{4}-){3}\d{4}|\d{16}')
    if not pattern.fullmatch(a) or count_consecutive(a.replace('-', ''))>= 4:
        return False
    else:
        return True

def adhar_card(a):
    pattern = re.compile(r'\d{4}\s\d{4}\s\d{4}')
    pattern2 = re.compile(r'\d{4}\d{4}\d{4}')
    if pattern.fullmatch(a) or pattern2.fullmatch(a):
        return True
    else:
        return False

def passport(a):
    pattern = re.compile(r'[A-Z]\d{2}\s\d{5}')
    pattern2 = re.compile(r'[A-Z]\d{2}\d{5}')
    if pattern.fullmatch(a) or pattern2.fullmatch(a):
        return True
    else:
        return False
def cheque_num(a):
    pattern = re.compile(r'\d{6}')
    if pattern.fullmatch(a):
        return True
    else:
        return False
def social_security_num(a):
    pattern = re.compile(r'^\d{3}-\d{2}-\d{4}$')
    pattern2 = re.compile(r'^\d{3}\d{2}\d{4}$')
    if pattern.fullmatch(a) or pattern2.fullmatch(a):
        return True
    else:
        return False
def check_request_method(a):
    if a=="POST" or a=="PUT" or a=="PATCH":
        return 1
    else:
        return 0

def check_third_party(a,domain):
    if a==None:
        return False
    if ("http://" in a or "https://" in a)  and (domain not in a):
        return True
    else:
        return False
def main():

    # parse xml file
    print('hi')

    host_global = input()
    parseXML('flipkart.xml')
    dc, cn, bc = "", "", ""
    outer, inner = 1, 0
    c1, c2, form = 0, 0, 0
    referer = ""
    common = {}


    for i in mydata:
        URL = i[0]
        # session,payment, method, referer,form
        if URL in common:
            max_s = max(i[-2], common[URL][0])
            max_p = max(i[-1], common[URL][1])
            max_form = max(i[6],common[URL][3])
        else:
            max_s, max_p,max_form = i[-2], i[-1],i[6]
        common[URL] = [max_s, max_p, i[2], max_form, i[3]]
        if URL in common:
            common[URL][3] += i[6]

    for i in mydata2:
        outer += 1
        URL = i[0]
        a = i[1]
        # session,payment, method, referer,form

        if URL in myglobal and a != None:
            list = myglobal[URL]
            list[0] += 1
        elif URL not in myglobal:
            if not a:
                myglobal[URL] = [0, 0]
            else:
                myglobal[URL] = [1, 0]
        else:
            print("")
        if not a:
            outer -= 1
            continue
        if a in myglobal:
            list = myglobal[a]
            list[1] += 1
        else:
            myglobal[a] = [0, 1]
        for j, k in enumerate(i):
            inner += 1
            if inner == 3:
                inner = 1
            ws2.cell(row=outer, column=inner).value = k

    for i in mylocation:
        Url = i[0]
        loc = i[1]
        for j in myglobal:
            if loc in j:
                G.add_edge(Url,j)
                if (Url,j) not in mydata2:
                    myglobal[Url][0]+=1
                    myglobal[j][1]+=1
                    ws2.cell(row = outer+1,column=1).value = Url
                    ws2.cell(row = outer+1,column=2).value = j
                    host1 = host_global
                    host2 = host_global
                    if Url in myhost:
                        host1 = myhost[Url]
                    if j in myhost:
                        host2 = myhost[Url]
                    ws2.cell(row=outer + 1, column=3).value = host1
                    ws2.cell(row=outer + 1, column=4).value = host2
                    outer+=1
    # print(myglobal)
    wbnew.save('Flipkart_new.xlsx')
    maxm = 2
    method = 0

    ics = nx.in_degree_centrality(G)
    # Gt = most_important(G) # trimming

    ocs = nx.out_degree_centrality(G)

    # Gt = most_important(G) # trimming

    cns = nx.closeness_centrality(G)

    pns = nx.betweenness_centrality(G)

    ens = nx.eigenvector_centrality_numpy(G)
    pgs = nx.pagerank(G, alpha=0.85, personalization=None,
                      max_iter=5, tol=1.0e-2, nstart=None, weight='weight',
                    dangling=None)

    def pagerank(G, alpha=0.85, personalization=None,
                 max_iter=5, tol=1.0e-2, nstart=None, weight='weight',
                 dangling=None):
        if len(G) == 0:
            return {}

        if not G.is_directed():
            D = G.to_directed()
        else:
            D = G

            # Create a copy in (right) stochastic form
        W = nx.stochastic_graph(D, weight=weight)
        N = W.number_of_nodes()

        nx.draw(D, with_labels=True, font_weight='bold')
        plt.show()

        # Choose fixed starting vector if not given
        if nstart is None:
            x = dict.fromkeys(W, 1.0 / N)
        else:
            # Normalized nstart vector
            s = float(sum(nstart.values()))
            x = dict((k, v / s) for k, v in nstart.items())

        if personalization is None:

            # Assign uniform personalization vector if not given
            p = dict.fromkeys(W, 1.0 / N)
        else:
            missing = set(G) - set(personalization)
            if missing:
                raise NetworkXError('Personalization dictionary '
                                    'must have a value for every node. '
                                    'Missing nodes %s' % missing)
            s = float(sum(personalization.values()))
            p = dict((k, v / s) for k, v in personalization.items())

        if dangling is None:

            # Use personalization vector if dangling vector not specified
            dangling_weights = p
        else:
            missing = set(G) - set(dangling)
            if missing:
                raise NetworkXError('Dangling node dictionary '
                                    'must have a value for every node. '
                                    'Missing nodes %s' % missing)
            s = float(sum(dangling.values()))
            dangling_weights = dict((k, v / s) for k, v in dangling.items())
        dangling_nodes = [n for n in W if W.out_degree(n, weight=weight) == 0.0]

        # power iteration: make up to max_iter iterations
        for _ in range(max_iter):
            xlast = x
            x = dict.fromkeys(xlast.keys(), 0)
            danglesum = alpha * sum(xlast[n] for n in dangling_nodes)
            for n in x:

                # this matrix multiply looks odd because it is
                # doing a left multiply x^T=xlast^T*W
                for nbr in W[n]:
                    x[nbr] += alpha * xlast[n] * W[n][nbr][weight]
                x[n] += danglesum * dangling_weights[n] + (1.0 - alpha) * p[n]

            # check convergence, l1 norm
            err = sum([abs(x[n] - xlast[n]) for n in x])
            if err < N * tol:
                return x
        raise NetworkXError('pagerank: power iteration failed to converge '
                            'in %d iterations.' % max_iter)



    for i in myglobal:
        c1, c2, form = 0, 0, 0
        method = 0
        referer = ""
        mythird = 0
        if i in common:
            c1 = common[i][0]
            c2 = common[i][1]
            method = check_request_method(common[i][2])
            form = common[i][3]
            mythird = third[i]
            host = common[i][4]
        else:
            c1 = string_text_processing(i, 'nlp')
            c2 = string_text_processing(i, 'session')
        ws3.cell(row=maxm,column=1).value = i
        l = myglobal[i]
        ws3.cell(row=maxm,column=2).value = l[0]
        ws3.cell(row=maxm,column=3).value = l[1]
        ws3.cell(row=maxm, column=4).value = host
        ws3.cell(row=maxm, column=5).value = ics[i]
        ws3.cell(row=maxm, column=6).value = ocs[i]
        ws3.cell(row=maxm, column=7).value = cns[i]
        ws3.cell(row=maxm, column=8).value = pns[i]
        ws3.cell(row=maxm, column=9).value = ens[i]
        ws3.cell(row=maxm, column=10).value = pgs[i]
        ws3.cell(row=maxm, column=11).value = c1
        ws3.cell(row=maxm, column=12).value = c2
        ws3.cell(row=maxm, column=13).value = form
        ws3.cell(row=maxm, column=14).value = method
        ws3.cell(row=maxm, column=15).value = mythird
        maxm+=1
    wbnew2.save('Flipkart_data.xlsx')
    nx.draw(G, with_labels=True)
    plt.show()

if __name__ == "__main__":
    # calling main function
    main()